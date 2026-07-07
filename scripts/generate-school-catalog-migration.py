from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\HP\Downloads\DMC691.xlsx")
OUTPUT = Path("supabase/migrations/20260707103000_add_school_catalog.sql")
CHUNK_SIZE = 500


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    if text in {"", "-", "None"}:
        return None
    return text


def sql_text(value):
    value = clean(value)
    if value is None:
        return "null"
    return "'" + value.replace("'", "''") + "'"


def sql_number(value):
    value = clean(value)
    if value is None:
        return "null"
    try:
        return repr(float(value))
    except ValueError:
        return "null"


def load_school_records():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    column = {name: index for index, name in enumerate(header)}
    records = []
    seen = set()

    for row in rows:
        if not row:
            continue
        smis = clean(row[column["smis"]])
        name = clean(row[column["ชื่อโรงเรียน"]])
        if not smis or not name or smis in seen:
            continue
        seen.add(smis)
        records.append(
            (
                sql_text(smis),
                sql_text(row[column["โรงเรียน"]]),
                sql_text(row[column["areacode"]]),
                sql_text(name),
                sql_text(row[column["ชื่ออำเภอ"]]),
                sql_text(row[column["ชื่อจังหวัด"]]),
                sql_text(row[column["ชื่อเขต"]]),
                sql_text(row[column["ประเภท_Text"]]),
                sql_text(row[column["ภาค_Text"]]),
                sql_text(row[column["ชั้นต่ำสุด_Text"]]),
                sql_text(row[column["ชั้นสูงสุด_Text"]]),
                sql_number(row[column["lat"]]),
                sql_number(row[column["long"]]),
            )
        )

    return records


def append_schema(lines):
    lines.extend(
        [
            "begin;",
            "",
            "create table if not exists public.school_catalog (",
            "  id text primary key,",
            "  school_code text unique,",
            "  area_code text,",
            "  name text not null,",
            "  district text,",
            "  province text,",
            "  education_area text,",
            "  school_type text,",
            "  region text,",
            "  lowest_grade text,",
            "  highest_grade text,",
            "  latitude double precision,",
            "  longitude double precision,",
            "  created_at timestamptz not null default now(),",
            "  updated_at timestamptz not null default now()",
            ");",
            "",
            "alter table public.school_catalog enable row level security;",
            "",
            'drop policy if exists "Anyone can search the school catalog" on public.school_catalog;',
            'create policy "Anyone can search the school catalog"',
            "on public.school_catalog",
            "for select",
            "to anon, authenticated",
            "using (true);",
            "",
            "grant select on public.school_catalog to anon, authenticated;",
            "",
            "create index if not exists school_catalog_name_idx on public.school_catalog (name);",
            "create index if not exists school_catalog_province_idx on public.school_catalog (province);",
            "create index if not exists school_catalog_education_area_idx on public.school_catalog (education_area);",
            "",
            "alter table public.profiles",
            "  add column if not exists school_id text references public.school_catalog(id) on delete set null;",
            "",
            "create index if not exists profiles_school_id_idx on public.profiles (school_id);",
            "",
            "grant update (school_id) on public.profiles to authenticated;",
            "",
        ]
    )


def append_school_data(lines, records):
    columns = (
        "(id, school_code, area_code, name, district, province, education_area, "
        "school_type, region, lowest_grade, highest_grade, latitude, longitude)"
    )

    for start in range(0, len(records), CHUNK_SIZE):
        chunk = records[start : start + CHUNK_SIZE]
        lines.append(f"insert into public.school_catalog {columns}")
        lines.append("values")
        for offset, record in enumerate(chunk):
            suffix = "," if offset < len(chunk) - 1 else ""
            lines.append("  (" + ", ".join(record) + ")" + suffix)
        lines.extend(
            [
                "on conflict (id) do update set",
                "  school_code = excluded.school_code,",
                "  area_code = excluded.area_code,",
                "  name = excluded.name,",
                "  district = excluded.district,",
                "  province = excluded.province,",
                "  education_area = excluded.education_area,",
                "  school_type = excluded.school_type,",
                "  region = excluded.region,",
                "  lowest_grade = excluded.lowest_grade,",
                "  highest_grade = excluded.highest_grade,",
                "  latitude = excluded.latitude,",
                "  longitude = excluded.longitude,",
                "  updated_at = now();",
                "",
            ]
        )


def append_profile_trigger(lines):
    lines.extend(
        [
            "-- New accounts remain students by default; school metadata is descriptive profile data only.",
            "create or replace function private.handle_new_user()",
            "returns trigger",
            "language plpgsql",
            "security definer",
            "set search_path = ''",
            "as $$",
            "declare",
            "  v_school_id text := nullif(btrim(new.raw_user_meta_data ->> 'school_id'), '');",
            "  v_school_name text := nullif(btrim(new.raw_user_meta_data ->> 'school_name'), '');",
            "begin",
            "  if v_school_id is not null then",
            "    select schools.id, schools.name",
            "      into v_school_id, v_school_name",
            "    from public.school_catalog as schools",
            "    where schools.id = v_school_id;",
            "",
            "    if not found then",
            "      v_school_id := null;",
            "      v_school_name := null;",
            "    end if;",
            "  elsif char_length(v_school_name) > 160 then",
            "    v_school_name := left(v_school_name, 160);",
            "  end if;",
            "",
            "  insert into public.profiles (id, role, display_name, email, school_id, school_name, total_points)",
            "  values (",
            "    new.id,",
            "    'student'::public.scisiam_user_role,",
            "    coalesce(",
            "      nullif(new.raw_user_meta_data ->> 'display_name', ''),",
            "      split_part(coalesce(new.email, ''), '@', 1),",
            "      'นักเรียน'",
            "    ),",
            "    new.email,",
            "    v_school_id,",
            "    v_school_name,",
            "    0",
            "  )",
            "  on conflict (id) do update set",
            "    email = excluded.email,",
            "    school_id = coalesce(excluded.school_id, public.profiles.school_id),",
            "    school_name = coalesce(excluded.school_name, public.profiles.school_name),",
            "    updated_at = now();",
            "",
            "  return new;",
            "end;",
            "$$;",
            "",
            "revoke all on function private.handle_new_user() from public;",
            "",
            "commit;",
            "",
        ]
    )


def main():
    records = load_school_records()
    lines = []
    append_schema(lines)
    append_school_data(lines, records)
    append_profile_trigger(lines)
    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUTPUT} with {len(records)} schools")


if __name__ == "__main__":
    main()
